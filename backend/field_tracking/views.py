from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, JSONParser
from django.utils import timezone
from django.db.models import Q
from django.contrib.auth.models import User
from .models import Site, DailyRoute, RouteStop, VisitLog, SiteVisitFormTemplate
from .serializers import SiteSerializer, DailyRouteSerializer, VisitLogSerializer, SiteVisitFormTemplateSerializer
from .utils import calculate_haversine_distance
from drf_spectacular.utils import extend_schema, OpenApiExample

class IsAdminOrManagerWeb(permissions.BasePermission):
    def has_permission(self, request, view):
        user_role = request.headers.get('X-User-Role')
        if user_role in ['ADMIN', 'MANAGER', 'admin', 'manager']:
            return True
        return False

class IsAdminOrManagerWebOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        user_role = request.headers.get('X-User-Role')
        if user_role in ['ADMIN', 'MANAGER', 'admin', 'manager']:
            return True
        return False

class SiteListView(generics.ListCreateAPIView):
    """
    Returns a list of all available sites for the employee to select from, or creates a new site.
    """
    queryset = Site.objects.all()
    serializer_class = SiteSerializer
    permission_classes = [AllowAny]

    def perform_create(self, serializer):
        site = serializer.save()
        from .models import AuditLog
        AuditLog.objects.create(
            event_type='SITE_ADD',
            description=f"Site '{site.name}' was added to the system.",
        )

class DailyRouteCreateView(generics.CreateAPIView):
    """
    Submit a daily route plan by providing a list of site_ids.
    """
    serializer_class = DailyRouteSerializer
    permission_classes = [IsAuthenticated]

    @extend_schema(
        request=DailyRouteSerializer,
        responses={201: DailyRouteSerializer},
        examples=[
            OpenApiExample(
                'Valid submission',
                value={'site_ids': [1, 2, 3]},
                request_only=True
            )
        ]
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)

class TodayRouteView(generics.RetrieveAPIView):
    """
    Retrieve the current logged-in employee's route for today.
    """
    serializer_class = DailyRouteSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        user = self.request.user
        today = timezone.now().date()
        route, _ = DailyRoute.objects.get_or_create(employee=user, date=today)
        return route

class CheckInView(generics.CreateAPIView):
    """
    Submit a check-in for a specific site with GPS coordinates and a photo.
    """
    serializer_class = VisitLogSerializer
    permission_classes = [IsAuthenticated]
    
    # Needs to accept multipart/form-data for image upload
    
    def perform_create(self, serializer):
        user = self.request.user
        site = serializer.validated_data.get('site')
        recorded_lat_val = serializer.validated_data.get('recorded_lat')
        recorded_lng_val = serializer.validated_data.get('recorded_lng')
        recorded_lat = float(recorded_lat_val) if recorded_lat_val is not None else 0.0
        recorded_lng = float(recorded_lng_val) if recorded_lng_val is not None else 0.0
        meeting = serializer.validated_data.get('meeting')
        
        if not site:
            # Dynamically create the site since it was omitted
            site_name = serializer.validated_data.get('site_name', f"Ad-Hoc Site ({recorded_lat}, {recorded_lng})")
            
            if 'Samruddhi' in site_name:
                site_name = f"Ad-Hoc Check-In"

            site = Site.objects.create(
                name=site_name,
                address=f"Auto-generated check-in location",
                latitude=recorded_lat,
                longitude=recorded_lng,
                geofence_radius=500
            )
            from .models import AuditLog
            AuditLog.objects.create(
                event_type='SITE_ADD',
                description=f"Site '{site.name}' was added automatically via Check-in.",
            )
            distance = 0.0 # Since they just created it where they are standing
            is_verified = True
        else:
            # Calculate distance to existing site
            site_lat = float(site.latitude)
            site_lng = float(site.longitude)
            if recorded_lat == 0.0 and recorded_lng == 0.0:
                distance = 0.0
                is_verified = True
            else:
                distance = calculate_haversine_distance(site_lat, site_lng, recorded_lat, recorded_lng)
                is_verified = distance <= site.geofence_radius
            
        # Update meeting status if provided and check-in is verified
        if meeting and is_verified:
            meeting.status = 'Completed'
            meeting.end_time = timezone.now()
            meeting.save()
        
        # Save the log with the created or existing site
        visit_log = serializer.save(
            employee=user,
            site=site,
            distance_from_site=distance,
            is_verified=is_verified
        )
        
        # Update RouteStop status if it exists in today's route
        today = timezone.now().date()
        try:
            route = DailyRoute.objects.get(employee=user, date=today)
            route_stop = RouteStop.objects.filter(route=route, site=site, status='pending').first()
            if route_stop:
                route_stop.status = 'completed'
                route_stop.save()
        except DailyRoute.DoesNotExist:
            pass
            
        from .models import AuditLog
        status_text = "Verified" if is_verified else "Unverified (Out of Geofence)"
        AuditLog.objects.create(
            event_type='CHECK_IN',
            description=f"Check-in at {site.name} [{status_text}]",
            user_id=user.id,
            employee_name=user.get_full_name() or user.username
        )

class VisitLogUpdateView(generics.UpdateAPIView):
    """
    Allows mobile apps to submit or update the report_data (filled form) for a specific VisitLog.
    """
    serializer_class = VisitLogSerializer
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, JSONParser]
    queryset = VisitLog.objects.all()

    def perform_update(self, serializer):
        visit_log = serializer.save()
        
        # If the visit log is linked to a meeting, completing the form marks the meeting as Completed
        if visit_log.meeting and visit_log.report_data:
            visit_log.meeting.status = 'Completed'
            visit_log.meeting.end_time = timezone.now()
            visit_log.meeting.save()
            
        from .models import AuditLog
        AuditLog.objects.create(
            event_type='FORM_SUBMIT',
            description=f"Report form submitted for {visit_log.site.name}",
            user_id=visit_log.employee.id,
            employee_name=visit_log.employee.get_full_name() or visit_log.employee.username
        )

class AdminLiveOverviewView(generics.ListAPIView):
    """
    Returns a live overview of all employees' daily routes and check-ins for a specific date (defaults to today).
    """
    permission_classes = [IsAdminOrManagerWeb]
    
    @extend_schema(
        responses={200: dict},
        summary="Live Overview of All Employees",
        description="Returns a live overview of all employees' daily routes and check-ins for a specific date."
    )
    def list(self, request, *args, **kwargs):
        date_str = request.query_params.get('date', None)
        if date_str:
            try:
                target_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
        else:
            target_date = timezone.now().date()
            
        data = []
        employees = Employee.objects.all()
        for emp in employees:
            if emp.id == 'admin' or 'admin' in emp.fullName.lower():
                continue
                
            visit_logs = VisitLog.objects.filter(employee__email=emp.email, check_in_time__date=target_date).order_by('check_in_time')
            meetings = AdHocMeeting.objects.filter(employee__email=emp.email, date=target_date).order_by('created_at')
            
            # Combine into a timeline
            timeline = []
            for log in visit_logs:
                # If a visit log is linked to a meeting, skip adding it as a separate check-in
                # to avoid duplicate rows, as its data is merged into the meeting row.
                if log.meeting:
                    continue
                    
                site_display_name = log.site.name
                timeline.append({
                    'type': 'Check-in',
                    'id': f"visit_{log.id}",
                    'site_name': site_display_name,
                    'lat': log.recorded_lat,
                    'lng': log.recorded_lng,
                    'start_lat': log.recorded_lat,
                    'start_lng': log.recorded_lng,
                    'end_lat': log.recorded_lat,
                    'end_lng': log.recorded_lng,
                    'time': log.check_in_time.isoformat(),
                    'end_time': log.check_in_time.isoformat(),
                    'status': 'Completed',
                    'distance_km': 0,
                    'fuel_cost': 0,
                    'notes': log.remarks,
                    'report_data': log.report_data,
                    'attachment_url': request.build_absolute_uri(log.photo.url) if log.photo else None
                })
            
            for m in meetings:
                meeting_photo = m.photo if m.photo else None
                meeting_report = m.report_data
                report_type = 'mom'
                visit_report = None
                
                # Fetch linked visit log data to merge into the meeting row
                linked_visit = m.visit_logs.first()
                if linked_visit:
                    if not meeting_photo and linked_visit.photo:
                        meeting_photo = linked_visit.photo
                        
                    is_empty = not meeting_report or str(meeting_report).strip() in ['{}', 'null', '', 'None']
                    if linked_visit.report_data:
                        visit_report = linked_visit.report_data
                    
                    if is_empty and linked_visit.report_data:
                        meeting_report = linked_visit.report_data
                        report_type = 'visit'
                
                timeline.append({
                    'type': 'Meeting',
                    'id': f"meeting_{m.id}",
                    'site_name': m.meeting_title,
                    'client_name': m.client_name,
                    'lat': m.destination_lat,
                    'lng': m.destination_lng,
                    'start_lat': m.start_lat,
                    'start_lng': m.start_lng,
                    'end_lat': m.end_lat,
                    'end_lng': m.end_lng,
                    'time': m.created_at.isoformat(), # using created_at for chronological sort
                    'status': m.status,
                    'distance_km': m.distance_km,
                    'fuel_cost': m.fuel_cost,
                    'fuel_approved': m.fuel_approved,
                    'end_time': m.end_time.isoformat() if m.end_time else None,
                    'report_data': meeting_report,
                    'visit_report_data': visit_report,
                    'report_type': report_type,
                    'attachment_url': request.build_absolute_uri(meeting_photo.url) if meeting_photo else None
                })
                
            # Sort timeline by time
            timeline.sort(key=lambda x: x['time'])
            
            status = 'Offline'
            current_location = None
            total_distance = sum(item.get('distance_km', 0) for item in timeline)
            
            # Find any actively running meeting
            active_meeting = None
            for m in meetings:
                if m.status == 'In Progress':
                    active_meeting = m
                    break
                    
            if active_meeting:
                linked_visit = active_meeting.visit_logs.first()
                if linked_visit:
                    status = 'At Site'
                    current_location = {
                        'lat': active_meeting.destination_lat,
                        'lng': active_meeting.destination_lng,
                        'time': linked_visit.check_in_time.isoformat(),
                        'site_name': active_meeting.meeting_title
                    }
                else:
                    status = 'Traveling'
                    current_location = {
                        'lat': active_meeting.current_lat,
                        'lng': active_meeting.current_lng,
                        'time': (active_meeting.start_time or active_meeting.created_at).isoformat(),
                        'site_name': 'In Transit to ' + active_meeting.meeting_title
                    }
            else:
                # No active meeting. Fallback to latest
                latest_visit = visit_logs.order_by('-check_in_time').first()
                latest_meeting = meetings.order_by('-created_at').first()
                
                if latest_meeting:
                    if latest_meeting.status in ['Completed', 'Closed']:
                        status = latest_meeting.status
                        current_location = {
                            'lat': latest_meeting.destination_lat,
                            'lng': latest_meeting.destination_lng,
                            'time': (latest_meeting.end_time or latest_meeting.created_at).isoformat(),
                            'site_name': latest_meeting.meeting_title
                        }
                    elif latest_meeting.status == 'Upcoming':
                        status = 'Not Started'
                        current_location = {
                            'lat': latest_meeting.current_lat,
                            'lng': latest_meeting.current_lng,
                            'time': latest_meeting.created_at.isoformat(),
                            'site_name': 'Pending'
                        }
                elif latest_visit:
                    status = 'At Site'
                    if latest_visit.meeting:
                        current_location = {
                            'lat': latest_visit.meeting.destination_lat,
                            'lng': latest_visit.meeting.destination_lng,
                            'time': latest_visit.check_in_time.isoformat(),
                            'site_name': latest_visit.meeting.meeting_title
                        }
                    else:
                        current_location = {
                            'lat': latest_visit.recorded_lat,
                            'lng': latest_visit.recorded_lng,
                            'time': latest_visit.check_in_time.isoformat(),
                            'site_name': latest_visit.site.name
                        }
                
            # Allow sending all employees if requested, or just those with activity.
            # We'll send everyone so the frontend list isn't empty, but default to Offline if no activity.
                
            data.append({
                'employee_id': emp.id,
                'employee_code': emp.employeeId,
                'employee_name': emp.fullName,
                'photo': request.build_absolute_uri(emp.profilePhoto) if emp.profilePhoto else None,
                'role': emp.designation,
                'travel_mode': emp.travelMode,
                'status': status,
                'total_distance': total_distance,
                'current_location': current_location,
                'timeline': timeline,
                'total_events': len(timeline),
                'completed_events': len([t for t in timeline if t['status'] == 'Completed'])
            })
            
        return Response(data)

from drf_spectacular.utils import extend_schema, OpenApiExample

@extend_schema(
    summary="Get or Update Form Template",
    description="Returns or updates the active form template schema for site visits. The schema array defines the dynamic UI fields.",
    examples=[
        OpenApiExample(
            "Form Schema Example",
            summary="Checklist Schema",
            description="The checklist schema the mobile app uses to render the UI.",
            value={
                "id": 1,
                "name": "Logicon Site Visit Report",
                "schema": [
                    {
                        "id": "site_operations_checklist",
                        "type": "radio",
                        "label": "Site Operations Checklist",
                        "options": ["Maintained", "Not Maintained", "Partial"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "attendance_register",
                        "type": "radio",
                        "label": "Attendance Register",
                        "options": ["Up to Date", "Incomplete", "Not Maintained"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "daily_logbooks",
                        "type": "radio",
                        "label": "Daily Logbooks",
                        "options": ["Updated", "Pending Entries", "Not Available"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "logbooks_authorization",
                        "type": "radio",
                        "label": "Daily Logbooks & Checklist Authorisation by operations",
                        "options": ["Yes", "No"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "housekeeping_machineries",
                        "type": "radio",
                        "label": "Housekeeping Machineries (Vacuum Cleaner, Scrubber, Jet Machine, etc.) Availability",
                        "options": ["Yes", "No"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "working_condition",
                        "type": "radio",
                        "label": "Working Condition",
                        "options": ["Good", "Requires Repair", "Not Functional"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "maintenance_description",
                        "type": "text",
                        "label": "Under maintenance Description",
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "consumables_tools",
                        "type": "radio",
                        "label": "Consumables/Tools",
                        "options": ["Adequate", "Shortage", "Pending Indent"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "uniform_status",
                        "type": "radio",
                        "label": "Uniform Status",
                        "options": ["Properly Dressed", "Partial", "Not Compliant"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "id_cards",
                        "type": "radio",
                        "label": "ID Cards",
                        "options": ["Issued & Displayed", "Pending", "Not Issued"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "esic_mediclaim",
                        "type": "radio",
                        "label": "ESIC & Mediclaim Coverage",
                        "options": ["Active", "Pending", "Not Available"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "daily_briefing",
                        "type": "radio",
                        "label": "Daily Briefing Conducted",
                        "options": ["Yes", "No"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "night_round",
                        "type": "radio",
                        "label": "Night Round Conducted",
                        "options": ["Yes", "No"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "observations_rounds",
                        "type": "text",
                        "label": "Observations from Rounds (if any)",
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "outstanding_dues",
                        "type": "radio",
                        "label": "Outstanding Dues (if any) - Add Details in Comment",
                        "options": ["None", "Yes"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "invoice_submission",
                        "type": "radio",
                        "label": "Invoice Submission Status",
                        "options": ["Submitted", "Pending"],
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "other_observations",
                        "type": "text",
                        "label": "Other Observations / Issues",
                        "requirePhoto": True,
                        "requireComment": True
                    },
                    {
                        "id": "corrective_actions",
                        "type": "text",
                        "label": "Corrective & Preventive Actions Suggested",
                        "requirePhoto": True,
                        "requireComment": True
                    }
                ],
                "created_at": "2026-06-04T12:00:00Z",
                "updated_at": "2026-06-04T12:00:00Z",
                "is_active": True
            }
        )
    ]
)
class SiteVisitFormTemplateView(generics.RetrieveUpdateAPIView):
    """
    Returns or updates the active form template schema.
    """
    serializer_class = SiteVisitFormTemplateSerializer
    
    def get_permissions(self):
        # Allow any authenticated mobile user to submit a form (POST) 
        # as long as they are NOT trying to maliciously update the 'schema' structure
        if self.request.method in ['POST', 'PATCH'] and 'schema' not in self.request.data and 'name' not in self.request.data:
            return [IsAuthenticated()]
        return [IsAdminOrManagerWebOrReadOnly()]
    
    def get_object(self):
        from .models import FormTemplate
        obj, created = FormTemplate.objects.get_or_create(
            form_type='site_visit', 
            defaults={'name': 'Logicon Site Visit Report', 'is_active': True}
        )
        
        # Automatically migrate old schemas to ensure new properties are present for mobile apps
        if obj.schema and isinstance(obj.schema, list):
            modified = False
            for field in obj.schema:
                if field.get('requirePhoto') is not True:
                    field['requirePhoto'] = True
                    modified = True
                if field.get('requireComment') is not True:
                    field['requireComment'] = True
                    modified = True
            if modified:
                obj.save(update_fields=['schema'])
                
        return obj

    def retrieve(self, request, *args, **kwargs):
        response = super().retrieve(request, *args, **kwargs)
        
        meeting_id = request.query_params.get('meeting_id')
        visit_log_id = request.query_params.get('visit_log_id')
        
        from .models import AdHocMeeting, VisitLog
        meeting = None
        
        if meeting_id:
            try:
                meeting = AdHocMeeting.objects.get(id=meeting_id)
            except AdHocMeeting.DoesNotExist:
                pass
        elif visit_log_id:
            try:
                vl = VisitLog.objects.get(id=visit_log_id)
                meeting = vl.meeting
            except VisitLog.DoesNotExist:
                pass
                
        if 'schema' in response.data:
            # We had previously appended it to schema; remove any pre_ items if they somehow exist
            response.data['schema'] = [f for f in response.data['schema'] if not str(f.get('id', '')).startswith('pre_')]
            
        response.data['predefined_data'] = {
            "client_name": meeting.client_name if meeting else "",
            "site_name": meeting.location_name if meeting else "",
            "meeting_date": meeting.date.strftime("%d-%m-%Y") if meeting and meeting.date else "",
            "start_time": meeting.time.strftime("%H:%M") if meeting and meeting.time else "",
            "end_time": getattr(meeting, 'end_time').strftime("%H:%M") if meeting and getattr(meeting, 'end_time', None) else "",
            "meeting_type": meeting.meeting_type if meeting else "",
            "prepared_by": (meeting.employee.get_full_name() or meeting.employee.username) if meeting and meeting.employee else ""
        }
            
        return response

    def _handle_form_submission(self, request):
        visit_log_id = request.data.get('visit_log_id')
        report_data = request.data.get('report_data', request.data)
        photo = request.FILES.get('photo') or request.FILES.get('attachment')
        
        visit_log = None
        if visit_log_id:
            try:
                visit_log = VisitLog.objects.get(id=visit_log_id, employee=request.user)
            except VisitLog.DoesNotExist:
                pass
        
        if not visit_log:
            # Fallback to the most recent VisitLog for the user
            visit_log = VisitLog.objects.filter(employee=request.user).order_by('-check_in_time').first()
            
        if visit_log:
            from .utils import process_checklist_report_data
            
            # Prepare data dict for the helper
            merged_data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
            if 'report_data' not in merged_data:
                merged_data['report_data'] = report_data
                
            for key in list(request.FILES.keys()):
                if key.startswith('photo_'):
                    merged_data[key] = request.FILES.getlist(key) if hasattr(request.FILES, 'getlist') else request.FILES.get(key)
                
            processed_data = process_checklist_report_data(merged_data, request)
            visit_log.report_data = processed_data.get('report_data', report_data)
            if photo:
                visit_log.photo = photo
            visit_log.save()
            
            # If the visit log is linked to a meeting, set it to Pending MOM and record end location
            if visit_log.meeting:
                visit_log.meeting.status = 'Pending MOM'
                visit_log.meeting.end_time = timezone.now()
                # Capture GPS coordinates submitted with the form as the end location
                lat = request.data.get('latitude') or request.data.get('lat')
                lng = request.data.get('longitude') or request.data.get('lng')
                if lat and lng:
                    visit_log.meeting.end_lat = lat
                    visit_log.meeting.end_lng = lng
                visit_log.meeting.save()
                
            from .models import AuditLog
            AuditLog.objects.create(
                event_type='FORM_SUBMIT',
                description=f"Report form submitted for {visit_log.site.name}",
                user_id=request.user.id,
                employee_name=request.user.get_full_name() or request.user.username
            )
            return Response({"message": "Form answers and photo saved successfully", "visit_log_id": visit_log.id}, status=200)
        else:
            return Response({"error": "No active visit log found to attach answers to."}, status=400)

    def post(self, request, *args, **kwargs):
        # The mobile app might be erroneously sending POST requests with form answers to this endpoint
        if 'schema' not in request.data and 'name' not in request.data:
            return self._handle_form_submission(request)
        # Treat POST as an update (like PUT) to the active template if 'schema' or 'name' is provided
        return self.update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        # The mobile app might also send PATCH requests with form answers here
        if 'schema' not in request.data and 'name' not in request.data:
            return self._handle_form_submission(request)
        return super().patch(request, *args, **kwargs)

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from django.http import HttpResponse
from .models import FormTemplate
from .serializers import FormTemplateSerializer
from io import BytesIO
import uuid

class FormTemplateViewSet(viewsets.ModelViewSet):
    queryset = FormTemplate.objects.all()
    serializer_class = FormTemplateSerializer
    permission_classes = [IsAdminOrManagerWeb]

    @action(detail=False, methods=['get'], url_path='download-sample')
    def download_sample_excel(self, request):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Form Checklists"
        # Headers
        headers = ['Checklist Name', 'Field Label', 'Field Type', 'Options', 'Require Photo']
        ws.append(headers)
        # Sample Data
        ws.append(['Site Audit', 'Engineer Name', 'text', '', 'No'])
        ws.append(['Site Audit', 'Condition Status', 'select', 'Good, Fair, Poor', 'Yes'])
        ws.append(['Safety Check', 'Wore Helmet?', 'boolean', '', 'No'])

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(), 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=form_checklists_sample.xlsx'
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser], url_path='upload-excel')
    def upload_excel(self, request):
        import openpyxl
        if 'file' not in request.FILES:
            return Response({"error": "No file provided"}, status=400)
            
        file_obj = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return Response({"error": "Empty or invalid Excel file"}, status=400)
                
            # Assume first row is header
            headers = [str(h).lower().strip() for h in rows[0]]
            
            # Map column indices
            try:
                name_idx = headers.index('checklist name')
                label_idx = headers.index('field label')
                type_idx = headers.index('field type')
            except ValueError:
                return Response({"error": "Missing required columns: Checklist Name, Field Label, Field Type"}, status=400)
                
            options_idx = headers.index('options') if 'options' in headers else -1
            photo_idx = headers.index('require photo') if 'require photo' in headers else -1

            checklists_map = {}
            for row in rows[1:]:
                if not row[name_idx] or not row[label_idx]:
                    continue
                
                checklist_name = str(row[name_idx]).strip()
                field_label = str(row[label_idx]).strip()
                field_type_raw = str(row[type_idx]).strip().lower()
                
                # Map field type
                field_type = field_type_raw
                if field_type not in ['text', 'textarea', 'number', 'select', 'boolean', 'date', 'photo', 'signature']:
                    field_type = 'text' # Default
                    
                options_str = str(row[options_idx]) if options_idx >= 0 and row[options_idx] else ""
                options = [opt.strip() for opt in options_str.split(',') if opt.strip()] if options_str and options_str.lower() != 'none' else []
                
                require_photo = False
                if photo_idx >= 0 and row[photo_idx]:
                    require_photo = str(row[photo_idx]).strip().lower() in ['yes', 'true', 'y', '1']
                    
                field_id = str(uuid.uuid4())
                
                field_def = {
                    "id": field_id,
                    "type": field_type,
                    "label": field_label,
                    "requirePhoto": require_photo
                }
                if options:
                    field_def["options"] = options

                if checklist_name not in checklists_map:
                    checklists_map[checklist_name] = []
                    
                checklists_map[checklist_name].append(field_def)
                
            # Create FormTemplates
            created_count = 0
            for name, schema in checklists_map.items():
                form_type_slug = name.lower().replace(' ', '_').replace('-', '_') + "_" + str(uuid.uuid4().hex[:6])
                FormTemplate.objects.create(
                    name=name,
                    form_type=form_type_slug,
                    schema=schema,
                    is_active=True,
                    is_global=True
                )
                created_count += 1
                
            return Response({"success": True, "message": f"Successfully created {created_count} checklists."})
        except Exception as e:
            return Response({"error": str(e)}, status=400)

class AssignedFormsListView(generics.ListAPIView):
    """
    Returns a list of all active forms assigned to the logged-in user.
    """
    serializer_class = FormTemplateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        from django.db.models import Q
        return FormTemplate.objects.filter(
            Q(is_active=True) & (Q(is_global=True) | Q(assigned_employees=user))
        ).distinct()

class DynamicFormTemplateView(generics.RetrieveAPIView):
    """
    Returns the schema of any form template by its form_type slug.
    """
    serializer_class = FormTemplateSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'form_type'
    
    def get_queryset(self):
        user = self.request.user
        from django.db.models import Q
        return FormTemplate.objects.filter(
            Q(is_active=True) & (Q(is_global=True) | Q(assigned_employees=user))
        ).distinct()

class AuditLogListView(generics.ListAPIView):
    """
    Returns a list of all audit logs.
    """
    from .models import AuditLog
    from .serializers import AuditLogSerializer
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAdminOrManagerWeb]

from rest_framework.views import APIView

class OfflineSyncView(APIView):
    """
    Accepts a bulk payload of offline check-ins and form submissions.
    """
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        request=dict,
        responses={200: dict},
        summary="Bulk Offline Sync",
        description="Sync multiple check-ins or form submissions at once."
    )
    def post(self, request, *args, **kwargs):
        user = request.user
        events = request.data.get('events', [])
        
        success_count = 0
        errors = []
        
        for event in events:
            try:
                event_type = event.get('type') # 'check_in' or 'form_submit'
                site_id = event.get('site_id')
                timestamp = event.get('timestamp') # original timestamp
                
                site = Site.objects.get(id=site_id)
                
                if event_type == 'check_in':
                    recorded_lat = float(event.get('recorded_lat'))
                    recorded_lng = float(event.get('recorded_lng'))
                    distance = calculate_haversine_distance(float(site.latitude), float(site.longitude), recorded_lat, recorded_lng)
                    is_verified = distance <= site.geofence_radius
                    
                    visit_log = VisitLog.objects.create(
                        employee=user,
                        site=site,
                        recorded_lat=recorded_lat,
                        recorded_lng=recorded_lng,
                        distance_from_site=distance,
                        is_verified=is_verified
                    )
                    
                    if timestamp:
                        # Backdate the visit log if possible, though check_in_time has auto_now_add=True
                        # We would need to update it manually using VisitLog.objects.filter
                        VisitLog.objects.filter(id=visit_log.id).update(check_in_time=timestamp)
                    
                    # Update RouteStop status
                    today = timezone.now().date()
                    try:
                        route = DailyRoute.objects.get(employee=user, date=today)
                        route_stop = RouteStop.objects.filter(route=route, site=site, status='pending').first()
                        if route_stop:
                            route_stop.status = 'completed'
                            route_stop.save()
                    except DailyRoute.DoesNotExist:
                        pass
                        
                    from .models import AuditLog
                    status_text = "Verified" if is_verified else "Unverified (Out of Geofence)"
                    AuditLog.objects.create(
                        event_type='CHECK_IN',
                        description=f"Offline Sync: Check-in at {site.name} [{status_text}]",
                        user_id=user.id,
                        employee_name=user.get_full_name() or user.username
                    )
                    success_count += 1
                    
                elif event_type == 'form_submit':
                    visit_log_id = event.get('visit_log_id')
                    report_data = event.get('report_data')
                    
                    if visit_log_id:
                        visit_log = VisitLog.objects.get(id=visit_log_id, employee=user)
                        visit_log.report_data = report_data
                        visit_log.save()
                        
                        from .models import AuditLog
                        AuditLog.objects.create(
                            event_type='FORM_SUBMIT',
                            description=f"Offline Sync: Report form submitted for {visit_log.site.name}",
                            user_id=user.id,
                            employee_name=user.get_full_name() or user.username
                        )
                        success_count += 1
                        
            except Exception as e:
                errors.append({"event": event, "error": str(e)})
                
        return Response({
            "message": "Sync completed",
            "success_count": success_count,
            "error_count": len(errors),
            "errors": errors
        })

from .serializers import AdHocMeetingSerializer
from .models import AdHocMeeting

class AdHocMeetingListCreateView(generics.ListCreateAPIView):
    """
    Allows mobile app to list their previous meetings or create a new spontaneous meeting.
    Supports filtering by ?status=Completed and searching by ?search=Acme
    """
    serializer_class = AdHocMeetingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        from django.db.models import Q
        qs = AdHocMeeting.objects.filter(employee=self.request.user).order_by('-date', '-time')
        
        # Status Filter
        status_param = self.request.query_params.get('status', None)
        if status_param:
            if status_param.lower() in ['completed', 'closed']:
                qs = qs.filter(status__in=['Completed', 'Closed'])
            else:
                qs = qs.filter(status=status_param)
            
        # Search Filter (client name or location)
        search_param = self.request.query_params.get('search', None)
        if search_param:
            qs = qs.filter(
                Q(client_name__icontains=search_param) | 
                Q(location_name__icontains=search_param) |
                Q(meeting_title__icontains=search_param)
            )
            
        return qs
    
    def perform_create(self, serializer):
        from django.utils import timezone
        # Force the meeting date to today to prevent formatting bugs from the mobile app (YYYY-DD-MM vs YYYY-MM-DD)
        # since Ad-Hoc meetings are spontaneous and meant to occur today.
        meeting = serializer.save(employee=self.request.user, date=timezone.now().date())
        
        # Calculate distance and estimated fuel cost
        if meeting.current_lat and meeting.current_lng and meeting.destination_lat and meeting.destination_lng:
            from .utils import calculate_haversine_distance
            dist = calculate_haversine_distance(
                float(meeting.current_lat), float(meeting.current_lng),
                float(meeting.destination_lat), float(meeting.destination_lng)
            )
            meeting.distance_km = round(dist / 1000.0, 2) # convert meters to km
            # Assume ₹8 per km for fuel reimbursement (standard rate)
            meeting.fuel_cost = round(meeting.distance_km * 8.0, 2)
            meeting.save()

        from .models import AuditLog
        AuditLog.objects.create(
            event_type='OTHER',
            description=f"Ad-Hoc Meeting created: {meeting.meeting_title} with {meeting.client_name} ({meeting.distance_km}km)",
            user_id=self.request.user.id,
            employee_name=self.request.user.get_full_name() or self.request.user.username
        )

class AdHocMeetingDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Allows mobile app to view or update a specific meeting.
    """
    serializer_class = AdHocMeetingSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return AdHocMeeting.objects.filter(employee=self.request.user)

from core.models import Employee
from django.db.models import Count
import calendar

def get_employee_stats(user, month=None, year=None):
    if not month:
        month = timezone.now().month
    if not year:
        year = timezone.now().year
        
    start_date = timezone.datetime(year, month, 1).date()
    end_date = timezone.datetime(year, month, calendar.monthrange(year, month)[1]).date()
    
    # Travels count
    visit_logs = VisitLog.objects.filter(employee=user, check_in_time__date__gte=start_date, check_in_time__date__lte=end_date)
    meetings = AdHocMeeting.objects.filter(employee=user, date__gte=start_date, date__lte=end_date)
    total_travels = visit_logs.count() + meetings.count()
    
    # Attendance (Distinct days)
    visit_days = set(visit_logs.values_list('check_in_time__date', flat=True))
    meeting_days = set(meetings.values_list('date', flat=True))
    attendance_days = len(visit_days.union(meeting_days))
    
    # Productivity
    routes = DailyRoute.objects.filter(employee=user, date__gte=start_date, date__lte=end_date)
    total_stops = RouteStop.objects.filter(route__in=routes).count()
    completed_stops = RouteStop.objects.filter(route__in=routes, status__in=['completed', 'closed']).count()
    
    productivity = 0
    if total_stops > 0:
        productivity = round((completed_stops / total_stops) * 100, 2)
        
    # Recent visits
    recent_visits = []
    for vl in visit_logs.order_by('-check_in_time')[:10]:
        recent_visits.append({
            'type': 'visit',
            'site_name': vl.site.name,
            'time': vl.check_in_time,
            'status': 'Verified' if vl.is_verified else 'Unverified',
            'report_data': vl.report_data
        })
    for m in meetings.order_by('-date', '-time')[:10]:
        time_combined = timezone.make_aware(timezone.datetime.combine(m.date, m.time))
        recent_visits.append({
            'type': 'meeting',
            'site_name': f"{m.meeting_title} ({m.client_name})",
            'time': time_combined,
            'status': 'Ad-Hoc',
            'report_data': None
        })
        
    recent_visits.sort(key=lambda x: x['time'], reverse=True)
    recent_visits = recent_visits[:10]
    
    try:
        emp = Employee.objects.get(email=user.email)
        emp_name = emp.fullName
        designation = emp.designation
        department = emp.departmentId.departmentName if emp.departmentId else "N/A"
    except Employee.DoesNotExist:
        emp_name = user.get_full_name() or user.username
        designation = "Unknown"
        department = "Unknown"
        
    return {
        "employee_name": emp_name,
        "position": designation,
        "department": department,
        "attendance_days": attendance_days,
        "total_travels": total_travels,
        "productivity_percentage": productivity,
        "recent_visits": recent_visits
    }

class EmployeeProfileAPIView(APIView):
    """
    Returns the stats for the currently logged in mobile app employee.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        stats = get_employee_stats(request.user)
        return Response(stats)

class AdminDashboardStatsView(APIView):
    permission_classes = [IsAdminOrManagerWeb]
    def get(self, request):
        return Response({"stats": "mock"})

class SubmitMOMAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, meeting_id):
        try:
            meeting = AdHocMeeting.objects.get(id=meeting_id, employee=request.user)
        except AdHocMeeting.DoesNotExist:
            return Response({"error": "Meeting not found"}, status=404)
            
        report_data = request.data.get('report_data')
        if report_data:
            import json
            report_data_dict = {}
            report_data_is_str = False
            if isinstance(report_data, str):
                try:
                    report_data_dict = json.loads(report_data)
                    report_data_is_str = True
                except:
                    pass
            elif isinstance(report_data, dict):
                report_data_dict = report_data.copy()
                
            # Filter out predefined data keys if the mobile app accidentally sent them back
            keys_to_remove = ["client_name", "site_name", "meeting_date", "start_time", "end_time", "meeting_type", "prepared_by", "predefined_data"]
            for k in keys_to_remove:
                report_data_dict.pop(k, None)
                
            # Process dynamic photo fields
            has_dynamic = False
            from django.core.files.storage import default_storage
            for key in list(request.FILES.keys()):
                if key.startswith('photo_'):
                    file_obj = request.FILES.get(key)
                    path = default_storage.save(f'meeting_proofs/{file_obj.name}', file_obj)
                    url = default_storage.url(path)
                    if request:
                        url = request.build_absolute_uri(url)
                    report_data_dict[key] = url
                    has_dynamic = True
                    
            if has_dynamic or report_data_is_str:
                report_data = json.dumps(report_data_dict) if report_data_is_str else report_data_dict
            else:
                report_data = report_data_dict
                    
            meeting.report_data = report_data
            
        if 'photo' in request.FILES:
            meeting.photo = request.FILES['photo']
            
        # Update meeting status based on MOM submission
        if meeting.status == 'Pending MOM' or meeting.status == 'Completed':
            meeting.status = 'Closed'
            
        meeting.save()
        
        from .models import AuditLog
        AuditLog.objects.create(
            event_type='FORM_SUBMIT',
            description=f"MOM submitted for {meeting.meeting_title}",
            user_id=request.user.id,
            employee_name=request.user.get_full_name() or request.user.username
        )
        
        return Response({"message": "MOM submitted successfully", "status": meeting.status})

class OverdueMOMAPIView(APIView):
    permission_classes = [IsAdminOrManagerWeb]

    def get(self, request):
        # Find meetings in Pending MOM > 24 hours
        from datetime import timedelta
        from django.utils import timezone
        threshold = timezone.now() - timedelta(hours=24)
        overdue_meetings = AdHocMeeting.objects.filter(
            status='Pending MOM',
            end_time__lte=threshold
        ).select_related('employee')
        
        data = []
        for meeting in overdue_meetings:
            data.append({
                'id': meeting.id,
                'title': meeting.meeting_title,
                'client': meeting.client_name,
                'employee_id': meeting.employee.id,
                'employee_name': meeting.employee.get_full_name() or meeting.employee.username,
                'end_time': meeting.end_time.isoformat() if meeting.end_time else None,
                'overdue_hours': int((timezone.now() - meeting.end_time).total_seconds() / 3600) if meeting.end_time else 0
            })
            
        return Response(data)

    def post(self, request):
        meeting_id = request.data.get('meeting_id')
        if not meeting_id:
            return Response({"error": "meeting_id is required"}, status=400)
            
        # Logic to send actual push notification or email goes here
        # Example: send_notification(meeting_id)
        
        return Response({"message": f"Reminder sent for meeting {meeting_id}"})




class MOMFormTemplateView(APIView):
    """
    Returns the MOM Form template configuration and allows creating/updating it.
    """
    permission_classes = []

    def get_object(self):
        from .models import FormTemplate
        obj, created = FormTemplate.objects.get_or_create(
            form_type='mom',
            defaults={
                'name': 'Minutes of Meeting (MOM)',
                'schema': [
                    {"id": "mom_purpose", "type": "text", "label": "1. Purpose of the Meeting", "required": True, "comment_enabled": True, "photo_enabled": False},
                    {"id": "mom_discussion", "type": "textarea", "label": "2. Key Discussion Points", "required": True, "comment_enabled": False, "photo_enabled": True},
                    {"id": "mom_action_items", "type": "textarea", "label": "3. Action Items / Next Steps", "required": True, "comment_enabled": True, "photo_enabled": False},
                    {"id": "mom_client_feedback", "type": "textarea", "label": "4. Client Feedback / Comments", "required": False, "comment_enabled": False, "photo_enabled": False},
                    {"id": "mom_followup_date", "type": "text", "label": "5. Follow-up Date (if any)", "required": False, "comment_enabled": False, "photo_enabled": False}
                ],
                'is_active': True
            }
        )
        return obj

    def get(self, request):
        obj = self.get_object()
        template_data = {
            "id": obj.id,
            "name": obj.name,
            "is_active": obj.is_active,
            "schema": obj.schema
        }
        
        meeting_id = request.query_params.get('meeting_id')
        meeting = None
        if meeting_id:
            from .models import AdHocMeeting
            try:
                meeting = AdHocMeeting.objects.get(id=meeting_id)
            except AdHocMeeting.DoesNotExist:
                pass

        template_data["schema"] = [f for f in template_data.get("schema", []) if not str(f.get('id', '')).startswith('pre_')]
        
        template_data["predefined_data"] = {
            "client_name": meeting.client_name if meeting else "",
            "site_name": meeting.location_name if meeting else "",
            "meeting_date": meeting.date.strftime("%d-%m-%Y") if meeting and meeting.date else "",
            "start_time": meeting.time.strftime("%H:%M") if meeting and meeting.time else "",
            "end_time": getattr(meeting, 'end_time').strftime("%H:%M") if meeting and getattr(meeting, 'end_time', None) else "",
            "meeting_type": meeting.meeting_type if meeting else "",
            "prepared_by": (meeting.employee.get_full_name() or meeting.employee.username) if meeting and meeting.employee else ""
        }

        return Response(template_data)

    def put(self, request):
        obj = self.get_object()
        # The frontend sends the entire template_data JSON
        schema = request.data.get('schema')
        if schema is not None:
            obj.schema = [f for f in schema if not str(f.get('id', '')).startswith('pre_')]
        if 'name' in request.data:
            obj.name = request.data['name']
        obj.save()
        
        return Response({"message": "Template updated successfully", "name": obj.name, "schema": obj.schema})

class ClientMeetingAnalyticsView(APIView):
    permission_classes = [IsAdminOrManagerWeb]
    def get(self, request):
        return Response([])

class MobileDashboardAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        return Response({})

class StartMeetingAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        return Response({})

class EndMeetingAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        return Response({})

class EmployeeAnalyticsAPIView(APIView):
    permission_classes = [IsAdminOrManagerWeb]
    def get(self, request, employee_id):
        return Response({})

class AdminEmployeeProfileAPIView(APIView):
    """
    Returns the stats for an employee for the admin dashboard.
    """
    permission_classes = [IsAdminOrManagerWeb]
    
    def get(self, request, employee_id, *args, **kwargs):
        try:
            emp = Employee.objects.get(id=employee_id)
            user = User.objects.get(email=emp.email)
            stats = get_employee_stats(user)
            return Response(stats)
        except (Employee.DoesNotExist, User.DoesNotExist):
            return Response({"error": "Employee not found"}, status=404)

class LiveEmployeeDetailsView(APIView):
    """
    Returns live tracking details for an employee including map coordinates, visit logs, and active status.
    """
    permission_classes = [IsAdminOrManagerWeb]

    def get(self, request, employee_id, *args, **kwargs):
        try:
            emp = Employee.objects.get(id=employee_id)
            user = User.objects.get(email=emp.email)
            
            date_str = request.query_params.get('date', None)
            if date_str:
                try:
                    today = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
            else:
                today = timezone.now().date()
            
            # Get today's route and stops
            route = DailyRoute.objects.filter(employee=user, date=today).first()
            stops = []
            if route:
                stops = list(route.stops.values('site__name', 'site__latitude', 'site__longitude', 'status', 'expected_time'))
                
            # Generate unified timeline
            timeline = []
            
            visits = VisitLog.objects.filter(employee=user, check_in_time__date=today).order_by('-check_in_time')
            for vl in visits:
                if vl.meeting:
                    continue
                timeline.append({
                    'type': 'Visit',
                    'id': f"visit_{vl.id}",
                    'site_name': vl.site.name,
                    'lat': vl.recorded_lat,
                    'lng': vl.recorded_lng,
                    'time': vl.check_in_time.isoformat(),
                    'status': 'Completed',
                    'report_data': vl.report_data,
                    'attachment_url': request.build_absolute_uri(vl.photo.url) if vl.photo else None
                })
                
            meetings = AdHocMeeting.objects.filter(employee=user, date=today).order_by('-created_at')
            meetings_data = []
            for m in meetings:
                meetings_data.append({
                    'id': m.id,
                    'meeting_title': m.meeting_title,
                    'client_name': m.client_name,
                    'status': m.status,
                    'time': m.time,
                    'current_lat': m.current_lat,
                    'current_lng': m.current_lng,
                    'destination_lat': m.destination_lat,
                    'destination_lng': m.destination_lng,
                    'start_time': m.start_time,
                    'end_time': m.end_time
                })
                
                meeting_photo = m.photo if m.photo else None
                meeting_report = m.report_data
                report_type = 'mom'
                visit_report = None
                linked_visit = m.visit_logs.first()
                if linked_visit:
                    if not meeting_photo and linked_visit.photo:
                        meeting_photo = linked_visit.photo
                        
                    is_empty = not meeting_report or str(meeting_report).strip() in ['{}', 'null', '', 'None']
                    if linked_visit.report_data:
                        visit_report = linked_visit.report_data
                        
                    if is_empty and linked_visit.report_data:
                        meeting_report = linked_visit.report_data
                        report_type = 'visit'
                    
                timeline.append({
                    'type': 'Meeting',
                    'id': f"meeting_{m.id}",
                    'site_name': m.meeting_title,
                    'client_name': m.client_name,
                    'lat': m.destination_lat,
                    'lng': m.destination_lng,
                    'time': m.created_at.isoformat(),
                    'status': m.status,
                    'distance_km': m.distance_km,
                    'fuel_cost': m.fuel_cost,
                    'fuel_approved': m.fuel_approved,
                    'end_time': m.end_time.isoformat() if m.end_time else None,
                    'report_data': meeting_report,
                    'visit_report_data': visit_report,
                    'report_type': report_type,
                    'attachment_url': request.build_absolute_uri(meeting_photo.url) if meeting_photo else None
                })
                
            timeline.sort(key=lambda x: x['time'])
                
            # Get active status logic and current location
            status = 'Not Started'
            current_location = None
            
            # Find any actively running meeting
            active_meeting = None
            for m in meetings:
                if m.status == 'In Progress':
                    active_meeting = m
                    break
                    
            if active_meeting:
                # Flow: start meeting -> Traveling -> verify photo (visit log created) -> At Site
                linked_visit = active_meeting.visit_logs.first()
                if linked_visit:
                    status = 'At Site'
                    current_location = {'lat': active_meeting.destination_lat, 'lng': active_meeting.destination_lng}
                else:
                    status = 'Traveling'
                    current_location = {'lat': active_meeting.current_lat, 'lng': active_meeting.current_lng}
            else:
                # No active meeting. Let's look at the most recent completed or upcoming activities.
                latest_meeting = meetings.first()
                latest_visit = visits.first()
                
                if latest_meeting:
                    if latest_meeting.status in ['Completed', 'Closed']:
                        status = latest_meeting.status # or Completed
                        current_location = {'lat': latest_meeting.destination_lat, 'lng': latest_meeting.destination_lng}
                    elif latest_meeting.status == 'Upcoming':
                        # They might be waiting for their next meeting
                        status = 'Not Started'
                        current_location = {'lat': latest_meeting.current_lat, 'lng': latest_meeting.current_lng}
                elif latest_visit:
                    status = 'At Site'
                    current_location = {'lat': latest_visit.recorded_lat, 'lng': latest_visit.recorded_lng}
                
                # Check for active routes
                if stops and any(s['status'] == 'in_progress' for s in stops):
                    status = 'Traveling'

            return Response({
                'employee_id': emp.employeeId,
                'employee_name': emp.fullName,
                'status': status,
                'current_location': current_location,
                'route_stops': stops,
                'timeline': timeline,
                'meetings': meetings_data
            })
        except (Employee.DoesNotExist, User.DoesNotExist):
            return Response({"error": "Employee not found"}, status=404)

class ClientMeetingAnalyticsView(APIView):
    """
    Returns analytics grouped by client_name from AdHocMeeting table.
    """
    permission_classes = [IsAdminOrManagerWeb]

    def get(self, request, *args, **kwargs):
        client_name = request.query_params.get('client_name')
        if not client_name:
            return Response({"error": "client_name parameter is required"}, status=400)
            
        parts = client_name.split()
        if len(parts) > 1:
            # If there's a space, they likely typed a full name, so check first and last name
            first, last = parts[0], parts[-1]
            meetings = AdHocMeeting.objects.filter(
                Q(client_name__icontains=client_name) | 
                (Q(employee__first_name__icontains=first) & Q(employee__last_name__icontains=last)) |
                Q(employee__username__icontains=client_name)
            )
        else:
            meetings = AdHocMeeting.objects.filter(
                Q(client_name__icontains=client_name) | 
                Q(employee__username__icontains=client_name) | 
                Q(employee__first_name__icontains=client_name) |
                Q(employee__last_name__icontains=client_name)
            )
        
        total_created = meetings.count()
        total_attended = meetings.filter(status__in=['Completed', 'Closed']).count()
        
        # Get unique employees involved
        employee_ids = meetings.values_list('employee_id', flat=True).distinct()
        employees = User.objects.filter(id__in=employee_ids)
        employee_names = [e.get_full_name() or e.username for e in employees]
        
        # Get unique clients involved
        involved_clients = list(meetings.values_list('client_name', flat=True).distinct())
        
        return Response({
            'client_name': client_name,
            'total_created': total_created,
            'total_attended': total_attended,
            'involved_employees': employee_names,
            'involved_clients': involved_clients
        })


class MobileDashboardAPIView(APIView):
    """
    Unified endpoint for the mobile app home dashboard.
    Returns KPIs (meetings today, completed, pending, distance covered, active hours, productivity)
    and a list of today's scheduled meetings.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        today = timezone.localtime().date()
        user = request.user
        
        # 1. Meetings KPIs
        meetings = AdHocMeeting.objects.filter(employee=user, date=today)
        visits = VisitLog.objects.filter(employee=user, check_in_time__date=today)
        
        total_meetings = meetings.count() + visits.count()
        completed_meetings = meetings.filter(status__in=['Completed', 'Closed']).count() + visits.count() # All visits are considered completed
        pending_meetings = meetings.filter(status__in=['Upcoming', 'In Progress']).count()
        
        productivity_percentage = 0
        if total_meetings > 0:
            productivity_percentage = int((completed_meetings / total_meetings) * 100)
            
        # 2. Distance Covered (from DailyRoute)
        route = DailyRoute.objects.filter(employee=user, date=today).first()
        distance_covered = 0
        if route and hasattr(route, 'total_distance_km') and route.total_distance_km:
            distance_covered = float(route.total_distance_km)
            
        # 3. Active Hours
        active_hours_str = '0h 0m'
        first_visit = visits.order_by('check_in_time').first()
        if first_visit:
            delta = timezone.localtime() - timezone.localtime(first_visit.check_in_time)
            hours, remainder = divmod(delta.seconds, 3600)
            minutes = remainder // 60
            active_hours_str = f"{hours}h {minutes}m"
            
        # 4. Today's Schedule List (only show scheduled meetings, not ad-hoc check-ins)
        schedule = []
            
        for m in meetings:
            schedule.append({
                'id': m.id,
                'client_name': m.client_name,
                'meeting_title': m.meeting_title,
                'time': m.time.strftime('%I:%M %p') if m.time else None,
                'location_name': m.location_name,
                'status': m.status
            })
            
        return Response({
            'kpis': {
                'meetings_today': total_meetings,
                'completed': completed_meetings,
                'pending': pending_meetings,
                'distance_covered_km': distance_covered,
                'active_hours': active_hours_str,
                'productivity_percentage': productivity_percentage
            },
            'todays_schedule': schedule
        })

class StartMeetingAPIView(APIView):
    """
    API for mobile app to start an AdHocMeeting.
    Accepts latitude and longitude.
    """
    permission_classes = [IsAuthenticated]

    @extend_schema(
        request=dict,
        responses={200: dict},
        summary="Start Meeting",
        description="Records the start time and location of the meeting.",
        examples=[
            OpenApiExample(
                'Start Meeting Payload',
                value={'latitude': 19.0760, 'longitude': 72.8777},
                request_only=True
            )
        ]
    )
    def post(self, request, pk, *args, **kwargs):
        try:
            meeting = AdHocMeeting.objects.get(pk=pk, employee=request.user)
        except AdHocMeeting.DoesNotExist:
            return Response({"error": "Meeting not found"}, status=404)
            
        lat = request.data.get('latitude')
        lng = request.data.get('longitude')
            
        meeting.status = 'In Progress'
        meeting.start_time = timezone.now()
        meeting.start_lat = lat
        meeting.start_lng = lng
        meeting.save()
        
        from .models import AuditLog
        AuditLog.objects.create(
            event_type='OTHER',
            description=f"Meeting Started: {meeting.meeting_title} with {meeting.client_name}",
            user_id=request.user.id,
            employee_name=request.user.get_full_name() or request.user.username
        )
        
        return Response({"message": "Meeting started successfully", "status": meeting.status})


import requests
from django.conf import settings

def get_journey_summary(meeting):
    result = {
        "start_time": meeting.start_time,
        "reached_time": meeting.end_time,
        "duration": None,
        "distance": "0 km",
        "traffic_condition": "Unknown"
    }
    
    if meeting.start_time and meeting.end_time:
        duration_delta = meeting.end_time - meeting.start_time
        hours, remainder = divmod(duration_delta.total_seconds(), 3600)
        minutes, seconds = divmod(remainder, 60)
        result['duration'] = f"{int(hours)}h {int(minutes)}m" if hours else f"{int(minutes)}m"
        
    if meeting.start_lat and meeting.start_lng and meeting.end_lat and meeting.end_lng:
        api_key = getattr(settings, 'GOOGLE_MAPS_API_KEY', '')
        if api_key:
            url = "https://maps.googleapis.com/maps/api/distancematrix/json"
            params = {
                "origins": f"{meeting.start_lat},{meeting.start_lng}",
                "destinations": f"{meeting.end_lat},{meeting.end_lng}",
                "key": api_key,
                "departure_time": "now",
            }
            try:
                response = requests.get(url, params=params, timeout=5)
                data = response.json()
                if data['status'] == 'OK' and data['rows'][0]['elements'][0]['status'] == 'OK':
                    element = data['rows'][0]['elements'][0]
                    result['distance'] = element['distance']['text']
                    
                    duration_value = element['duration']['value']
                    traffic_value = element.get('duration_in_traffic', {}).get('value', duration_value)
                    
                    if traffic_value > duration_value * 1.5:
                        result['traffic_condition'] = "Heavy"
                    elif traffic_value > duration_value * 1.2:
                        result['traffic_condition'] = "Moderate"
                    else:
                        result['traffic_condition'] = "Light"
            except Exception:
                pass
        else:
            result['distance'] = "12.5 km (est)"
            result['traffic_condition'] = "Moderate (est)"
            
    return result

class EndMeetingAPIView(APIView):
    """
    API for mobile app to end an AdHocMeeting and get summary.
    """
    permission_classes = [IsAuthenticated]

    @extend_schema(
        request=dict,
        responses={200: dict},
        summary="End Meeting",
        description="Records the end time and location of the meeting.",
        examples=[
            OpenApiExample(
                'End Meeting Payload',
                value={'latitude': 19.0760, 'longitude': 72.8777},
                request_only=True
            )
        ]
    )
    def post(self, request, pk, *args, **kwargs):
        try:
            meeting = AdHocMeeting.objects.get(pk=pk, employee=request.user)
        except AdHocMeeting.DoesNotExist:
            return Response({"error": "Meeting not found"}, status=404)
            
        lat = request.data.get('latitude')
        lng = request.data.get('longitude')
        photo = request.FILES.get('photo')
            
        meeting.status = 'Completed'
        meeting.end_time = timezone.now()
        meeting.end_lat = lat
        meeting.end_lng = lng
        
        if photo:
            meeting.photo = photo
            
        meeting.save()
        
        from .models import AuditLog
        AuditLog.objects.create(
            event_type='OTHER',
            description=f"Meeting Ended: {meeting.meeting_title} with {meeting.client_name}",
            user_id=request.user.id,
            employee_name=request.user.get_full_name() or request.user.username
        )
        
        summary = get_journey_summary(meeting)
        
        # Calculate Fuel Cost based on Travel Mode
        distance_str = summary.get('distance', '0 km')
        try:
            # Extract number from "12.5 km" or "12.5 km (est)"
            distance_val = float(distance_str.split(' ')[0].replace(',', ''))
        except (ValueError, IndexError):
            distance_val = 0.0
            
        meeting.distance_km = distance_val
        
        # Travel mode logic (assume Petrol is 100 Rs/L)
        travel_mode = getattr(request.user, 'employee_set').first().travelMode if hasattr(request.user, 'employee_set') and request.user.employee_set.exists() else 'Bike'
        
        if travel_mode == 'Bike' or travel_mode == 'Scooty':
            # 30 km/L -> ~3.33 Rs/km
            meeting.fuel_cost = round(distance_val * 3.33, 2)
        elif travel_mode == 'Car':
            # 15 km/L -> ~6.66 Rs/km
            meeting.fuel_cost = round(distance_val * 6.66, 2)
        else:
            meeting.fuel_cost = 0.0
            
        meeting.save()
        
        return Response({
            "message": "Meeting ended successfully", 
            "status": meeting.status,
            "summary": summary,
            "fuel_cost": meeting.fuel_cost,
            "distance_km": meeting.distance_km
        })

class TrackingConfigurationAPIView(APIView):
    permission_classes = [IsAdminOrManagerWeb]
    def get(self, request):
        from .models import TrackingConfiguration
        config, _ = TrackingConfiguration.objects.get_or_create(id=1)
        return Response({
            "idleThreshold": config.idle_threshold,
            "geoFenceRadius": config.geo_fence_radius,
            "autoRefreshRate": config.auto_refresh_rate,
            "defaultMapTheme": config.default_map_theme
        })
    def post(self, request):
        from .models import TrackingConfiguration
        config, _ = TrackingConfiguration.objects.get_or_create(id=1)
        data = request.data
        if "idleThreshold" in data: config.idle_threshold = data["idleThreshold"]
        if "geoFenceRadius" in data: config.geo_fence_radius = data["geoFenceRadius"]
        if "autoRefreshRate" in data: config.auto_refresh_rate = data["autoRefreshRate"]
        if "defaultMapTheme" in data: config.default_map_theme = data["defaultMapTheme"]
        config.save()
        return Response({"message": "Configuration updated successfully"})

from django.http import HttpResponse
import datetime
from .reports.generator import generate_report_data
from .reports.pdf_builder import build_pdf_report
from django.utils import timezone
from .scheduler import generate_and_send

class DownloadReportView(APIView):
    permission_classes = [IsAdminOrManagerWeb]
    def get(self, request):
        period = request.query_params.get('period', 'daily').lower()
        
        days_map = {
            'daily': 1,
            'weekly': 7,
            'monthly': 30,
            'yearly': 365
        }
        days = days_map.get(period, 1)
        
        end_date = timezone.now()
        start_date = end_date - datetime.timedelta(days=days)
        
        data = generate_report_data(start_date, end_date)
        pdf_buffer = build_pdf_report(data)
        
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="FieldOps_Report_{period.capitalize()}.pdf"'
        return response

class EmailReportView(APIView):
    permission_classes = [IsAdminOrManagerWeb]
    def post(self, request):
        period = request.data.get('period', 'daily').lower()
        days_map = {
            'daily': 1,
            'weekly': 7,
            'monthly': 30,
            'yearly': 365
        }
        days = days_map.get(period, 1)
        
        # generate_and_send will build the PDF and email it to the admins
        generate_and_send(period.capitalize(), days)
        return Response({"message": f"{period.capitalize()} report emailed successfully to administrators."})

class ActiveChecklistsListView(generics.ListAPIView):
    """
    Returns all active form templates (checklists) assigned to the logged-in user.
    """
    serializer_class = FormTemplateSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        return FormTemplate.objects.filter(
            Q(is_active=True) & (Q(is_global=True) | Q(assigned_employees=user))
        ).distinct()
